import sys
import os
import pandas as pd
import numpy as np

from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation

# import SetRefinedSheet
import warnings
warnings.simplefilter("ignore", UserWarning)
# Ignorer les warnings spécifiques d'openpyxl
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl.styles.stylesheet")

def main():
    print("\n Exécution du fichier : SetCities.py")
    if len(sys.argv) < 3:
        print("Erreur : Deux arguments sont attendus (nom du fichier et département).")
        sys.exit(1)

    inputFilePath = "./cities/cities.xlsx"
    output_filepath = sys.argv[1]
    department = sys.argv[2]

    if department.startswith("0"):
        department = department[1:]


    try:
        if not os.path.exists(inputFilePath):
            print(f"Erreur : Le fichier source '{inputFilePath}' est introuvable.")
            sys.exit(1)

        # Charger le fichier source avec pandas
        df = pd.read_excel(inputFilePath, sheet_name=None)
        if "Sheet1" not in df:
            print("Erreur : La feuille 'Sheet1' est introuvable dans le fichier source.")
            sys.exit(1)

        df_cities = df["Sheet1"]
        first_column = df_cities.columns[0]
        df_cities[first_column] = pd.to_numeric(df_cities[first_column], errors='coerce').fillna(0).astype(int)



        # Vérifier que la première colonne contient des valeurs
        if df_cities[first_column].isna().all():
            print("Erreur : La colonne des départements ne contient aucune donnée valide.")
            sys.exit(1)

        # Filtrer les lignes correspondant au département
        df_filtered = df_cities[df_cities[first_column].astype(str).str.match(rf"^{department}\d{{3}}$", na=False)]
        # Vérifier si le fichier de sortie existe
        if not os.path.exists(output_filepath):
            print(f"Erreur : Le fichier de sortie '{output_filepath}' est introuvable.")
            sys.exit(1)

        # Charger le fichier Excel avec openpyxl pour préserver les formules
        wb = load_workbook(output_filepath)

        # Vérifier les feuilles disponibles
        sheet_names = [sheet.strip() for sheet in wb.sheetnames]  # Supprimer espaces invisibles

        if "cities" not in sheet_names:
            print("Erreur : La feuille 'cities' est introuvable dans le fichier de destination.")
            sys.exit(1)

        ws = wb["cities"]

        # Effacer uniquement les valeurs des colonnes A:J sans toucher aux formules
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=10):
            for cell in row:
                if cell.data_type != 'f':  # Ne pas toucher aux formules
                    cell.value = None

        # Réécrire les nouvelles valeurs à partir de la ligne 1
        for r_idx, row in enumerate(df_filtered.itertuples(index=False), start=2):
            for c_idx, value in enumerate(row[:10], start=1):  # Écriture sur les colonnes A:J
                ws.cell(row=r_idx, column=c_idx, value=value)

        # Sauvegarde du fichier Excel
        wb.save(output_filepath)
        wb.close()

    except FileNotFoundError as e:
        print(f"Erreur : {e}")
        sys.exit(1)
    except Exception as e:
        print(f"Erreur inattendue : {e}")
        sys.exit(1)

    
    # Traitement de la feuille "Commodites"
    wb = load_workbook(output_filepath)
    if "Commodites" not in wb.sheetnames:
        wb.create_sheet(title="Commodites")

    ws = wb["Commodites"]

    # Initialiser les chaînes de caractères pour les colonnes C et E
    colCAmenities = ""
    colEAmenities = ""
    cellJ1Profitability = ""

    # Parcourir les lignes de la colonne C et E
    for i in range(3, 1000):
        cell_CAmenities = ws.cell(row=i, column=3)
        cell_EAmenities = ws.cell(row=i, column=5)
        cell_J1Profitability = ws.cell(row=i, column=4)

        if cell_CAmenities.value is not None:
            colCAmenities += str(cell_CAmenities.value) + ","

        if cell_EAmenities.value is not None:
            colEAmenities += str(cell_EAmenities.value) + ","

        if cell_J1Profitability.value is not None:
            cellJ1Profitability += str(cell_J1Profitability.value) + ","

    # Supprimer la dernière virgule
    colCAmenities = colCAmenities.rstrip(',')
    colEAmenities = colEAmenities.rstrip(',')
    cellJ1Profitability = cellJ1Profitability.rstrip(',')

    # Ajouter les validations de données
    dvcAmenities = DataValidation(type="list", formula1="=Reference!$M:$M", showDropDown=False)
    dveAmenities = DataValidation(type="list", formula1="=Reference!$E:$E", showDropDown=False)
    dvJ1Amenities = DataValidation(type="list", formula1="=Reference!$M:$M", showDropDown=False)

    ws.add_data_validation(dvcAmenities)
    ws.add_data_validation(dveAmenities)

    dvcAmenities.add("A3:A35")
    dveAmenities.add("B2:W2")

    # Ajouter la formule en anglais de B3 à W95
    for row in range(3, 96):
        for col in range(2, 24):  # B=2, W=23
            cell = ws.cell(row=row, column=col)
            cell.value = ('=IFERROR(SUMIFS(INDIRECT(_xlfn.XLOOKUP({0}2,Reference!$E:$E,Reference!$I:$I)), '
                          'INDIRECT(_xlfn.XLOOKUP({0}2,Reference!$E:$E,Reference!$H:$H)),Commodites!$A{1}), "")'
                          .format(chr(64 + col), row))

    ws = wb["Rentabilite"]
    ws.add_data_validation(dvJ1Amenities)
    dvJ1Amenities.add("N1")
    
    ref_ws = wb["Reference"]
    ws["N1"] = ref_ws["M2"].value



    wb.save(output_filepath)
    wb.close()

if __name__ == "__main__":
    main()
