import os
import warnings
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.datavalidation import DataValidation
import pandas as pd
import sys
import warnings

warnings.simplefilter("ignore", UserWarning)

# Ignorer les warnings spécifiques d'openpyxl
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl.styles.stylesheet")
if len(sys.argv) < 3:
    print("Erreur : Deux arguments sont attendus (nom du fichier et département).")
    sys.exit(1)  # Quitte le programme avec une erreur

output_filename = sys.argv[1]
department = sys.argv[2]




# Dictionnaire de correspondance entre les noms des fichiers et les noms des feuilles souhaités
sheet_name_mapping = {
    "Commerces": "Commerces",
    "Enseignements": "Enseignements",
    "Sante": "Santé",
    "Services": "Services",
    "SportLoisirCulture": "SportLoisirCulture",
    "Tourisme": "Tourisme",
    "Transport": "Transport"
}


def filterDepartement(department):
    folder_path = "./commodites"
    files = os.listdir(folder_path)
    excel_files = [f for f in files if f.endswith('.xlsx')]

    if os.path.exists(output_filename):
        book = load_workbook(output_filename)
    else:
        book = Workbook()
        if 'Sheet' in book.sheetnames:
            del book['Sheet']

    for file in excel_files:
        file_path = os.path.join(folder_path, file)
        try:
            df = pd.read_excel(file_path, sheet_name="COM")
            header_df = df.head(7)
            filtered_df = df[df.iloc[:, 0].astype(str).str.startswith(department)]
            result_df = pd.concat([header_df, filtered_df], ignore_index=True)

            if "COMMERCE" in file:
                sheet_name = "Commerces"
            elif "ENSEIGNEMENT" in file:
                sheet_name = "Enseignements"
            elif "SANTE_ACTION_SOCIALE" in file:
                sheet_name = "Sante"
            elif "SERVICE_PARTICULIER" in file:
                sheet_name = "Services"
            elif "SPORT_LOISIR_CULTURE" in file:
                sheet_name = "SportLoisirCulture"
            elif "TOURISME" in file:
                sheet_name = "Tourisme"
            elif "TRANSPORT_DEPLACEMENT" in file:
                sheet_name = "Transport"
            else:
                print(f"{file} non traité")
                continue

            # Si la feuille existe déjà, effacer son contenu
            if sheet_name in book.sheetnames:
                sheet = book[sheet_name]
                for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
                    for cell in row:
                        cell.value = None  # Efface uniquement le contenu des cellules
            else:
                sheet = book.create_sheet(title=sheet_name)

            # Écrire le DataFrame résultant dans la feuille en commençant à la ligne 1
            sheet.delete_rows(1, sheet.max_row)  # Supprime toutes les lignes existantes
            sheet.sheet_state = "hidden"
            for r in dataframe_to_rows(result_df, index=False, header=True):
                sheet.append(r)

        except ValueError as e:
            print(f"Erreur lors de la lecture du fichier {file}: {e}")

    # Sauvegarder le fichier après avoir traité tous les fichiers
    book.save(output_filename)

    # Traitement de la feuille "Commodites"
    wb = load_workbook(output_filename)
    if "Commodites" not in wb.sheetnames:
        wb.create_sheet(title="Commodites")

    ws = wb["Commodites"]

    # Initialiser les chaînes de caractères pour les colonnes C et E
    colCAmenities = ""
    colEAmenities = ""
    cellJ1Profitability = ""

    # Parcourir les lignes de la colonne C et E
    for i in range(3, 1000):
        cell_CAmenities = ws.cell(row=i, column=3)
        cell_EAmenities = ws.cell(row=i, column=5)
        cell_J1Profitability = ws.cell(row=i, column=4)

        if cell_CAmenities.value is not None:
            colCAmenities += str(cell_CAmenities.value) + ","

        if cell_EAmenities.value is not None:
            colEAmenities += str(cell_EAmenities.value) + ","

        if cell_J1Profitability.value is not None:
            cellJ1Profitability += str(cell_J1Profitability.value) + ","

    # Supprimer la dernière virgule
    colCAmenities = colCAmenities.rstrip(',')
    colEAmenities = colEAmenities.rstrip(',')
    cellJ1Profitability = cellJ1Profitability.rstrip(',')

    # Ajouter les validations de données
    dvcAmenities = DataValidation(type="list", formula1="=Reference!$M:$M", showDropDown=False)
    dveAmenities = DataValidation(type="list", formula1="=Reference!$E:$E", showDropDown=False)
    dvJ1Amenities = DataValidation(type="list", formula1="=Reference!$M:$M", showDropDown=False)

    ws.add_data_validation(dvcAmenities)
    ws.add_data_validation(dveAmenities)

    dvcAmenities.add("A3:A35")
    dveAmenities.add("B2:W2")

    # Ajouter la formule en anglais de B3 à W95
    for row in range(3, 96):
        for col in range(2, 24):  # B=2, W=23
            cell = ws.cell(row=row, column=col)
            cell.value = ('=IFERROR(SUMIFS(INDIRECT(_xlfn.XLOOKUP({0}2,Reference!$E:$E,Reference!$I:$I)), '
                          'INDIRECT(_xlfn.XLOOKUP({0}2,Reference!$E:$E,Reference!$H:$H)),Commodites!$A{1}), "")'
                          .format(chr(64 + col), row))

    ws = wb["Rentabilite"]
    ws.add_data_validation(dvJ1Amenities)
    dvJ1Amenities.add("N1")
    
    ref_ws = wb["Reference"]
    ws["N1"] = ref_ws["M2"].value



    wb.save(output_filename)
    wb.close()

def main():
    print("\n Exécution du fichier : SetAmenitiesSheets.py")
    filterDepartement(department)

if __name__ == "__main__":
    main()
