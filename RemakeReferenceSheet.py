import pandas as pd
import openpyxl
import os
import unicodedata
import re
import sys

# Ignorer les warnings
import warnings
warnings.simplefilter("ignore", UserWarning)
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl.styles.stylesheet")

print("\n Exécution du fichier : RemakeReferenceSheet.py")

# Nom du fichier Excel
fichier_sortie = sys.argv[1]

# Vérifier si le fichier existe déjà
if os.path.exists(fichier_sortie):
    wb = openpyxl.load_workbook(fichier_sortie)
else:
    print("Le fichier n'existe pas.")
    exit()

# Vérifier si les feuilles nécessaires existent
if "Commerces" not in wb.sheetnames or "Reference" not in wb.sheetnames:
    print("Feuilles manquantes dans le fichier.")
    exit()

# Charger les feuilles nécessaires
ws_commerces = wb["Commerces"]
ws_reference = wb["Reference"]

# Effacer l'ancien contenu des colonnes M et N
for row in ws_reference.iter_rows(min_row=2, min_col=13, max_col=14):
    for cell in row:
        cell.value = None

# Récupérer les valeurs uniques de la colonne B à partir de la ligne 9
donnees_commerces = set()
for row in ws_commerces.iter_rows(min_row=9, min_col=2, max_col=2, values_only=True):
    if row[0]:
        donnees_commerces.add(row[0])

donnees_commerces = sorted(donnees_commerces)  # Trier pour plus de lisibilité

# Fonction pour nettoyer les noms
def nettoyer_nom(nom):
    nom = ''.join(c for c in unicodedata.normalize('NFD', nom) if unicodedata.category(c) != 'Mn')
    nom = nom.replace("-", " ")
    nom = re.sub(r"\bSaint\b", "st", nom, flags=re.IGNORECASE)
    return nom

# Ajouter les valeurs dans les colonnes M et N de la feuille Reference
for i, valeur in enumerate(donnees_commerces, start=2):
    ws_reference[f"M{i}"] = valeur
    ws_reference[f"N{i}"] = nettoyer_nom(valeur)

# Vérifier si la feuille 'Refined' existe
if "Refined" not in wb.sheetnames:
    print("La feuille 'Refined' est manquante.")
    exit()

ws_refined = wb["Refined"]

# Effacer le contenu des colonnes V et W de la feuille Reference
for row in ws_reference.iter_rows(min_row=2, min_col=22, max_col=23):
    for cell in row:
        cell.value = None

# Récupérer les valeurs uniques de la colonne D de la feuille Refined
donnees_refined = set()
for row in ws_refined.iter_rows(min_row=2, min_col=4, max_col=4, values_only=True):
    if row[0]:
        donnees_refined.add(row[0])

donnees_refined = sorted(donnees_refined)  # Trier pour plus de lisibilité

# Ajouter les valeurs dans les colonnes V et W de la feuille Reference
for i, valeur in enumerate(donnees_refined, start=2):
    ws_reference[f"V{i}"] = valeur
    ws_reference[f"W{i}"] = nettoyer_nom(valeur)

# Sauvegarder le fichier modifié
wb.save(fichier_sortie)
