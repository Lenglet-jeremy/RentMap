import sys
import openpyxl

if len(sys.argv) < 2:
    print("Usage: python script.py <fichier_excel>")
    sys.exit(1)

sortieExcel = sys.argv[1]

# Extraire la partie avant le dernier tiret et normaliser en minuscules
split_val = sortieExcel.rsplit("-", 1)[0].casefold()

# Charger le fichier Excel et la feuille "Refined"
wb = openpyxl.load_workbook(sortieExcel)
if "Refined" not in wb.sheetnames:
    print("La feuille 'Refined' n'existe pas dans le fichier.")
    sys.exit(1)

ws = wb["Refined"]

# Parcours en sens inverse
for row in range(ws.max_row, 1, -1):  # De la dernière ligne jusqu'à la 2e
    cell_value = ws[f"C{row}"].value
    if cell_value and str(cell_value).casefold() != split_val:
        ws.delete_rows(row)

# Sauvegarder le fichier Excel
wb.save(sortieExcel)
wb.close()

print(f"Nettoyage terminé. Les lignes ne correspondant pas à '{split_val}' ont été supprimées.")
