from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl import load_workbook
from bs4 import BeautifulSoup
from datetime import datetime
from openpyxl.worksheet.datavalidation import DataValidation
import unicodedata

import requests
import time
import random
import os
import re
import sys

import warnings
warnings.simplefilter("ignore", UserWarning)
# Ignorer les warnings spécifiques d'openpyxl
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl.styles.stylesheet")

if len(sys.argv) >= 3 : 
    output_filepath = sys.argv[1]
    sortieExcel = sys.argv[1]
    department = sys.argv[2]
    AdministraveRegionURL = sys.argv[3]

BaseURL = "https://www.meilleursagents.com"
DepartmentURLS = []
CitysLink = []
CitysURL = {}

# Chemin du dossierSource et du fichier Excel
dossierSource = f"./data/neighborhood/{sortieExcel.split(".")[0]}"
feuille_cible = "Refined"

rejected = ["rue-", "allee-", "avenue-", "boulevard-", "chemin-", "impasse-", "route-", "lotissement-", "place-", "residence-", "pont-"]

headers = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:118.0) Gecko/20100101 Firefox/118.0",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.5",
    "Accept-Encoding": "gzip, deflate, br",
    "Connection": "keep-alive",
    "Upgrade-Insecure-Requests": "1",
    "Sec-Fetch-Dest": "document",
    "Sec-Fetch-Mode": "navigate",
    "Sec-Fetch-Site": "none",
    "Sec-Fetch-User": "?1"
}

import unicodedata

def normalize_text(text):
    if text is None:
        return None
    # Transformer en minuscule et enlever les accents
    normalized_text = unicodedata.normalize('NFD', text.casefold())
    normalized_text = ''.join(c for c in normalized_text if unicodedata.category(c) != 'Mn')
    # Remplacer espaces et apostrophes par des tirets
    normalized_text = normalized_text.replace(" ", "-").replace("'", "-")
    
    print(f"Returned Value : {normalized_text}")
    return normalized_text


def timeSleep() : 
    # sleepTime = random.uniform(60, 180)
    sleepTime = random.uniform(2, 5)
    print(f"Pause de {sleepTime} secondes")
    time.sleep(sleepTime)

def charger_ou_creer_excel(nom_fichier, nom_feuille):
    """Charge le fichier Excel s'il existe, sinon le crée et retourne la feuille cible."""
    entetes = [
        "Adresse", "Département", "Localité", "Ville", "Quartier",
        "Prix au m² appartement", "Prix au m² maison",
        "Loyer au m² appartement", "Loyer au m² maison", "Ville Renommé",
        "Date MAJ MeilleursAgents", "DateMAJ Rentmap", "lien"
    ]

    if os.path.exists(nom_fichier):
        wb = load_workbook(nom_fichier)
        if nom_feuille not in wb.sheetnames:
            ws = wb.create_sheet(nom_feuille)
        else:
            ws = wb[nom_feuille]
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = nom_feuille

    ws.append(entetes)  # Ajouter les en-têtes si la feuille est vide

    return wb, ws

def extract_data_for_type(type_bien, biens, prices_and_rents):
    """Extrait les informations spécifiques pour chaque type de bien (appartement, maison)."""
    for bien in biens:
        bien_type = bien.find('p', class_='text--small').get_text(strip=True)
        
        # Récupérer le prix ou loyer
        price_range = bien.find_next('ul', class_='prices-summary__price-range')
        price_type = price_range.find('li', class_='text--small').get_text(strip=True)
        price_value = price_range.find('li', class_='big-number').get_text(strip=True)

        # Nettoyage de la valeur
        price_value = price_value.replace(' ', '').replace('€', '').replace(',', '.').strip()

        # Assigner en fonction du type de bien et du type de prix/loyer
        if bien_type == "Appartement":
            if "Prix m² moyen" in price_type:
                prices_and_rents['prix_m2_appartement'] = price_value
            elif "Loyer mensuel/m² moyen" in price_type:
                prices_and_rents['loyer_m2_appartement'] = price_value
        elif bien_type == "Maison":
            if "Prix m² moyen" in price_type:
                prices_and_rents['prix_m2_maison'] = price_value
            elif "Loyer mensuel/m² moyen" in price_type:
                prices_and_rents['loyer_m2_maison'] = price_value

    
    return prices_and_rents


def extract_canonical_href(html):
    soup = BeautifulSoup(html, 'html.parser')
    link_tag = soup.find('link', {'rel': 'canonical'})
    if link_tag and 'href' in link_tag.attrs:
        return link_tag['href']
    return None

def checkpath(dossierSource):
    if not os.path.exists(dossierSource):
        os.makedirs(dossierSource)

def extractData(dossierSource, feuille, wb): 
    """Parcourt les fichiers HTML et extrait les informations de localisation et de prix/loyer."""
    checkpath(dossierSource)
    for i, fichier in enumerate(os.listdir(dossierSource)):
        cheminFichier = os.path.join(dossierSource, fichier)
        
        if os.path.isfile(cheminFichier) and fichier.endswith(".html"):
            with open(cheminFichier, "r", encoding="utf-8") as f:
                contenu = f.read()
                soup = BeautifulSoup(contenu, "html.parser")

            lien = extract_canonical_href(contenu)

            # Extraction de l'adresse
            adresseTitle = soup.find("h1", {"data-prices-sell-title": True, "class": "prices-summary__title"})
            adresse = adresseTitle.text.strip() if adresseTitle else None

            # Localisation
            pointerSpans = soup.find_all("span", {"itemprop": "name", "class": "pointer-event"})
            split_val = sortieExcel.rsplit("-", 1)[0]
            print("="*50)
            text1 = normalize_text(pointerSpans[2].text.strip()) if len(pointerSpans) >= 3 else None
            text2 = normalize_text(split_val)
            print("="*50)

            if adresse and "Prix immobilier " in adresse and text1 == text2: 
            # if True:
                departement = pointerSpans[1].text.strip() if len(pointerSpans) >= 2 else None
                localite = pointerSpans[2].text.strip() if len(pointerSpans) >= 3 else None
                ville = pointerSpans[3].text.strip() if len(pointerSpans) >= 4 else None
                quartier = pointerSpans[4].text.strip() if len(pointerSpans) >= 5 else None

                # Extraction de la date d'estimation
                time_tag = soup.find("time")
                date_estimation = time_tag["datetime"] if time_tag else None

                # Initialisation des variables pour les prix
                prix_m2_appartement, prix_m2_maison = None, None
                loyer_m2_appartement, loyer_m2_maison = None, None

                # Extraction des prix et loyers des appartements
                appartement_sections = soup.find_all('div', class_='prices-summary__apartment-prices')
                for section in appartement_sections:
                    price_items = section.find('ul', class_='prices-summary__price-range').find_all('li')
                    for j in range(len(price_items) - 1):
                        label = price_items[j].get_text(strip=True)
                        value = price_items[j + 1].get_text(strip=True).replace(' ', '').replace('€', '').replace(',', '.').strip()

                        if "Prix m2 moyen" in label:
                            prix_m2_appartement = int(value)
                        elif "Loyer mensuel/m2 moyen" in label:
                            loyer_m2_appartement = float(value)

                # Extraction des prix et loyers des maisons
                maison_sections = soup.find_all('div', class_='prices-summary__house-prices')
                for section in maison_sections:
                    price_items = section.find('ul', class_='prices-summary__price-range').find_all('li')
                    for j in range(len(price_items) - 1):
                        label = price_items[j].get_text(strip=True)
                        value = price_items[j + 1].get_text(strip=True).replace(' ', '').replace('€', '').replace(',', '.').strip()

                        if "Prix m2 moyen" in label:
                            prix_m2_maison = int(value)
                        elif "Loyer mensuel/m2 moyen" in label:
                            loyer_m2_maison = float(value)

                # Date de mise à jour
                date_heure_courante = datetime.now()
                DateMajRentMap = date_heure_courante.strftime("%Y-%m-%d %H:%M:%S")

                # Trouver la ligne où insérer les données
                row_index = feuille.max_row + 1

                # Ajout des données dans la feuille
                feuille.append([adresse, departement, localite, ville, quartier, prix_m2_appartement, prix_m2_maison, loyer_m2_appartement, loyer_m2_maison, "", date_estimation, DateMajRentMap, lien])

                # Ajout manuel de la formule dans la cellule
                feuille.cell(row=row_index, column=10).value = f"=_xlfn.XLOOKUP(D{row_index},Reference!V:V,Reference!W:W)"
    
    wb.save(sortieExcel)

        

def sauvegarder_excel(wb, nom_fichier):
    """Enregistre le fichier Excel."""

    # Charger le fichier Excel pour ajouter les validations de données
    wb = load_workbook(nom_fichier)

    # Vérifier si la feuille "Commodités" existe
    if "Commodités" not in wb.sheetnames:
        wb.create_sheet(title="Commodites")

    ws = wb["Commodites"]

    # Initialiser les chaînes de caractères pour les colonnes C et E
    colCAmenities = ""
    colEAmenities = ""
    cellJ1Profitability = ""

    # Parcourir les lignes de la colonne C et E
    for i in range(3, 1000):
        cell_CAmenities = ws.cell(row=i, column=3)
        cell_EAmenities = ws.cell(row=i, column=5)
        cell_J1Profitability = ws.cell(row= i, column=4)

        if cell_CAmenities.value is not None:
            colCAmenities += str(cell_CAmenities.value) + ","

        if cell_EAmenities.value is not None:
            colEAmenities += str(cell_EAmenities.value) + ","

        if cell_J1Profitability.value is not None:
            cellJ1Profitability += str(cell_J1Profitability.value) + ","

    # Supprimer la dernière virgule
    colCAmenities = colCAmenities.rstrip(',')
    colEAmenities = colEAmenities.rstrip(',')
    cellJ1Profitability = cellJ1Profitability.rstrip(',')

    # Ajouter les validations de données
    dvcAmenities = DataValidation(type="list", formula1="=Reference!$M:$M", showDropDown=False)
    dveAmenities = DataValidation(type="list", formula1="=Reference!$E:$E", showDropDown=False)
    dvJ1Amenities = DataValidation(type="list", formula1="=Reference!$M:$M", showDropDown=False)

    ws.add_data_validation(dvcAmenities)
    ws.add_data_validation(dveAmenities)

    dvcAmenities.add("A3:A35")
    dveAmenities.add("B2:W2")

    
    # Ajouter la formule en anglais de B3 à W95
    for row in range(3, 96):
        for col in range(2, 24):  # B=2, W=23
            cell = ws.cell(row=row, column=col)
            cell.value = ('=IFERROR(SUMIFS(INDIRECT(_xlfn.XLOOKUP({0}2,Reference!$E:$E,Reference!$I:$I)), '
                        'INDIRECT(_xlfn.XLOOKUP({0}2,Reference!$E:$E,Reference!$H:$H)),Commodites!$A{1}), "")'
                        .format(chr(64 + col), row))
            

    ws = wb["Rentabilite"]
    ws.add_data_validation(dvJ1Amenities)
    dvJ1Amenities.add("J1")

    wb.save(nom_fichier)
    wb.close()


def effacer_contenu_feuille(feuille, wb):
    """Efface tout le contenu de la feuille spécifiée."""
    feuille.delete_rows(2, feuille.max_row)
    wb.save(sortieExcel)


# Scrape administrative region to get department links
def scrapeAdministrativeRegionURL(administraveRegionURL=""): 
    
    try:
        timeSleep()

        # Request
        response = requests.get(administraveRegionURL, headers=headers)
        response.raise_for_status()
        
        dataDir = f"./data/administrativeRegions/{administraveRegionURL.split("/")[4]}"
        if not os.path.exists(dataDir):
            os.makedirs(dataDir)
            print(f"Répertoire créé : {dataDir}")

        # Save to file
        fileName = f"{administraveRegionURL.split("/")[4]}.html"
        with open(f"{dataDir}/{fileName}", 'w', encoding="utf-8") as file : 
            file.write(response.text)

        extractAdministrativeRegionFile(f"{dataDir}/{fileName}")
    
    except Exception as e:
        print(f"Une erreur inattendue est survenue : {e}")
    
# Extract administrative region file to get all departement link
def extractAdministrativeRegionFile(filePath):
    with open(filePath, 'r', encoding='utf-8') as file:
        htmlContent = file.read()

    soup = BeautifulSoup(htmlContent, 'html.parser')
    links = soup.find_all('a', {'data-ua-event-action': 'go:prices:subregion'})
    
    for link in links:
        href = link.get('href')
        if href:
            DepartmentURLS.append(f"{BaseURL}{href}")

    scrapeDepartementURLs(DepartmentURLS)

# Scrape department URL to get all citys
def scrapeDepartementURLs(departmentURLS):
    try:
        # creating output folder
        dataDir = "./data/departements"
        if not os.path.exists(dataDir):
            os.makedirs(dataDir)
            print(f"Répertoire créé : {dataDir}")

        # Request and save response to file
        for i, departmentURL in enumerate(departmentURLS):
            timeSleep()

            response = requests.get(departmentURL, headers=headers)

            # Vérification de l'état de la réponse
            if response.status_code == 200:
                fileName = departmentURL.split("/")[-2] + ".html"
                filePath = os.path.join(dataDir, fileName)

                # Enregistrer le contenu HTML dans un fichier
                with open(filePath, 'w', encoding="utf-8") as file:
                    file.write(response.text)
                

                extractDepartmentFiles(filePath)
            else:
                print(f"Erreur lors de la requête pour {departmentURL}: Statut {response.status_code}")

    except requests.exceptions.RequestException as e:
        print(f"Erreur réseau ou requête HTTP : {e}")

# Extract department file to get all citys link
def extractDepartmentFiles(filePath):
    with open(filePath, 'r', encoding='utf-8') as file: 
        htmlContent = file.read()

    soup = BeautifulSoup(htmlContent, 'html.parser')
    links = soup.find_all('a', class_="btn btn--secondary btn--small")
    
    for link in links:
        href = link.get('href')
        if href:
            CitysLink.append(f"{BaseURL}{href}")
            scrapeCitysLink(f"{BaseURL}{href}")

def scrapeCitysLink(link): 
    try:
        timeSleep()

        # Effectuer une requête HTTP pour récupérer le contenu de la page
        response = requests.get(link, headers=headers)
        response.raise_for_status()  # Vérifie si la requête a réussi

        # Déterminer le nom du fichier et créer le répertoire si nécessaire
        cityDir = "./data/citysLinks"
        if not os.path.exists(cityDir):
            os.makedirs(cityDir)

        cityFileName = f"{link.split('/')[4]}_Citys.html"
        cityFilePath = os.path.join(cityDir, cityFileName)

        # Enregistrer le contenu HTML dans un fichier
        with open(cityFilePath, 'w', encoding="utf-8") as file:
            file.write(response.text)

        extractCitysLink(cityFilePath)

    except requests.exceptions.RequestException as e:
        print(f"Erreur réseau ou requête HTTP : {e}")
    except Exception as e:
        print(f"Une erreur inattendue est survenue : {e}")

def extractCitysLink(filePath):
    # Extraire la localisation à partir du chemin du fichier
    location = filePath.split('\\')[1]
    location = location.split('_')[0]
    
    try:
        # Lire le contenu HTML du fichier
        with open(filePath, 'r', encoding='utf-8') as file:
            htmlContent = file.read()

        # Parser le contenu avec BeautifulSoup
        soup = BeautifulSoup(htmlContent, 'html.parser')

        # Trouver toutes les balises <a> avec l'attribut data-medium spécifié
        cityLinks = soup.find_all('a', {"data-medium": "price_map_subregion_cities_index_city_links_price_map_city"})

        # Si la localisation n'existe pas encore dans le dictionnaire, initialisez-la avec une liste vide
        if location not in CitysURL:
            CitysURL[location] = []

        # Ajouter les liens à la liste associée à la localisation
        for link in cityLinks:
            href = link.get('href')
            if href:
                fullLink = f"{BaseURL}{href}"
                CitysURL[location].append(fullLink)  # Ajouter le lien au dictionnaire
                scrapeCity(href, location)

    except Exception as e:
        print(f"Erreur lors de l'extraction des liens de villes depuis {filePath} : {e}")

def scrapeCity(link, location):
    try:
        # Déterminer le nom du fichier à partir du lien
        fileName = f"{link.split('/')[2]}.html"  # Exemple : 'ville-name.html'

        # Créer le répertoire pour stocker les fichiers
        cityDir = f"./data/citys/{location}"
        if not os.path.exists(cityDir):
            os.makedirs(cityDir)

        # Chemin complet du fichier
        filePath = os.path.join(cityDir, fileName)

        if not os.path.exists(filePath) : 
            # Pause aléatoire pour éviter d'être détecté comme bot
            timeSleep()

            # Effectuer une requête HTTP pour récupérer le contenu de la page
            response = requests.get(f"{BaseURL}{link}", headers=headers)
            response.raise_for_status()  # Vérifie si la requête a réussi

            # Enregistrer le contenu HTML dans un fichier
            with open(filePath, 'w', encoding='utf-8') as file:
                file.write(response.text)

        # Extraire les données de la ville
        extractCity(filePath, location)

    except requests.exceptions.RequestException as e:
        print(f"Erreur réseau ou requête HTTP pour {link} : {e}")
    except Exception as e:
        print(f"Une erreur inattendue est survenue pour {link} : {e}")

def extractCity(filePath, location):
    """
    Extract city details from the given HTML file and scrape its neighborhoods.
    """
    try:
        with open(filePath, 'r', encoding='utf-8') as file:
            htmlContent = file.read()

        soup = BeautifulSoup(htmlContent, 'html.parser')

        # Extraction des quartiers
        neighborhoodLinks = soup.find_all('a', href=re.compile(r'^/prix-immobilier/.*'))
        neighborhoods = []

        for link in neighborhoodLinks:
            href = link.get('href')
            if href:
                neighborhoods.append(f"{BaseURL}{href}")


        # Scraper chaque quartier
        for link in neighborhoods:
            scrapeNeighborhood(link, location)

    except Exception as e:
        print(f"Erreur lors de l'extraction des données de la ville depuis {filePath} : {e}")

def scrapeAllCities():
    """
    Scrape all cities and their neighborhoods for the given administrative region.
    """
    for department in DepartmentURLS:
        scrapeDepartementURLs([department])  # Traite chaque département
        for city in CitysLink:
            scrapeCity(city, department)


def scrapeNeighborhood(link, location):
    """
    Scrape data for a specific neighborhood and save to file if it doesn't already exist.
    """

    if any(keyword in link for keyword in rejected):
        return

    try:
        # Déterminer le répertoire de stockage
        neighborhoodDir = f"./data/neighborhood/{location}"
        if not os.path.exists(neighborhoodDir):
            os.makedirs(neighborhoodDir)

        # Nommer le fichier en fonction du lien
        neighborhoodName = link.split("/")[-2]
        filePath = os.path.join(neighborhoodDir, f"{neighborhoodName}.html")

        # Vérifier si le fichier existe déjà
        if os.path.exists(filePath):
            # Obtenir la date de modification du fichier
            file_mod_time = datetime.fromtimestamp(os.path.getmtime(filePath))
            today = datetime.today()

            # Vérifier si le fichier date du mois dernier et si nous sommes au moins le 2 du mois courant
            if file_mod_time.month == (today.month - 1) % 12 and today.day >= 2:
                print(f"Fichier {filePath} périmé. Relancement du scraping.")
            else:
                print(f"Fichier déjà présent : {filePath}. Passage au prochain lien.")
                return
        else:
            print(f"Le fichier {filePath} n'existe pas. Lancement du scraping.")

        # Pause aléatoire pour éviter la détection comme bot
        timeSleep()

        # Requête pour récupérer le contenu
        response = requests.get(link, headers=headers)
        response.raise_for_status()

        # Enregistrer le contenu HTML dans un fichier
        with open(filePath, 'w', encoding='utf-8') as file:
            file.write(response.text)

        # Extraire les données du fichier
        extractNeighborhood(filePath)

    except requests.exceptions.RequestException as e:
        print(f"Erreur réseau ou requête HTTP pour {link} : {e}")
    except Exception as e:
        print(f"Une erreur inattendue est survenue pour {link} : {e}")
        
# ===========================================================
# Refaire cette fonction pour presenter les donnée dans Excel
# ===========================================================


def extractNeighborhood(filePath):
    """
    Extract neighborhood details from the given HTML file and append them to an Excel file.
    """
    try:
        # Charger le fichier HTML
        with open(filePath, 'r', encoding='utf-8') as file:
            htmlContent = file.read()

        soup = BeautifulSoup(htmlContent, 'html.parser')

        # Extraire les informations nécessaires
        # Exemple : Titre du quartier, prix moyen au m²
        neighborhoodTitle = soup.find("h1", class_="title").get_text(strip=True) if soup.find("h1", class_="title") else "Inconnu"
        pricePerSquareMeter = soup.find("div", class_="averagePrice").get_text(strip=True) if soup.find("div", class_="averagePrice") else "Inconnu"
        rentPerSquareMeter = soup.find("div", class_="averageRent").get_text(strip=True) if soup.find("div", class_="averageRent") else "Inconnu"

        # Ajouter ces données à un fichier Excel
        appendToExcel([
            neighborhoodTitle,
            pricePerSquareMeter,
            rentPerSquareMeter,
            filePath
        ])

    except Exception as e:
        print(f"Erreur lors de l'extraction des données du fichier {filePath} : {e}")


def appendToExcel(data):
    """
    Append the given data as a new row in the Excel file.
    """
    excelFile = "./data/NeighborhoodData.xlsx"
    if not os.path.exists(excelFile):
        # Créer un nouveau fichier Excel si non existant
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Données Quartiers"
        # Ajouter les en-têtes
        sheet.append(["Nom du Quartier", "Prix au m²", "Loyer au m²", "Chemin du fichier"])
        workbook.save(excelFile)

    # Charger le fichier existant et ajouter une ligne
    workbook = load_workbook(excelFile)
    sheet = workbook.active
    sheet.append(data)
    workbook.save(excelFile)



def main():
    print("Execution du fichier  : SetRefinedSheet.py")
    # 1. Scraper la région administrative
    # scrapeAdministrativeRegionURL(AdministraveRegionURL)

    # 2. Extraire les liens des départements
    # for department in DepartmentURLS:
    #     scrapeDepartementURLs([department])

    # 3. Scraper les villes et quartiers
    # scrapeAllCities()

    # scrapeCity("/prix-immobilier/rennes-35000/", "rennes-35000")

    # 4. Extraire les données et les enregistrer dans Excel
    wb, feuille_bdd = charger_ou_creer_excel(sortieExcel, feuille_cible)
    effacer_contenu_feuille(feuille_bdd, wb)
    extractData(dossierSource, feuille_bdd, wb)
    sauvegarder_excel(wb, sortieExcel)
    pass



if __name__ == "__main__":
    main()
