import pandas as pd
import sys
import os

from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.datavalidation import DataValidation

# import SetRefinedSheet
import warnings
warnings.simplefilter("ignore", UserWarning)
# Ignorer les warnings spécifiques d'openpyxl
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl.styles.stylesheet")

def main():
    print("\n Exécution du fichier : SetPopulationSheet.py")
    if len(sys.argv) < 3:
        print("Erreur : Deux arguments sont attendus (nom du fichier et département).")
        sys.exit(1)

    output_file = sys.argv[1]
    department = str(sys.argv[2])
    input_file = "./PopulationIRIS/base-cc-evol-struct-pop-2021.CSV"
    

    # Charger le fichier CSV sans séparateur spécifique (détecte automatiquement)
    df = pd.read_csv(input_file, sep=";", dtype=str)

    # Filtrer les lignes où la colonne A commence par la valeur du département
    filtered_df = df[df.iloc[:, 0].astype(str).str.startswith(department)].iloc[:, :9]





    # Vérifier si le fichier Excel existe
    if os.path.exists(output_file):
        try:
            book = load_workbook(output_file)
            if "Population" in book.sheetnames:
                sheet = book["Population"]
                # Effacer le contenu des colonnes A à I (à partir de la 2e ligne)
                for row in sheet.iter_rows(min_col=1, max_col=9, min_row=2):
                    for cell in row:
                        cell.value = None
            else:
                sheet = book.create_sheet("Population")
        except Exception as e:
            print(f"Erreur lors de l'ouverture du fichier Excel : {e}")
            print("Création d'un nouveau fichier Excel.")
            book = Workbook()
            sheet = book.active
            sheet.title = "Population"
    else:
        # Créer un nouveau fichier Excel
        book = Workbook()
        sheet = book.active
        sheet.title = "Population"

    # Écrire les données dans les colonnes A à I
    for row_idx, row in enumerate(filtered_df.itertuples(index=False), start=2):
        for col_idx, value in enumerate(row, start=1):
            sheet.cell(row=row_idx, column=col_idx, value=value)
        for row in sheet.iter_rows(min_col=1, max_col=2):
            for cell in row:
                if cell.value is not None:  # Vérifier que la cellule n'est pas vide
                    try:
                        cell.value = int(cell.value)  # Convertir en entier
                    except ValueError:
                        pass  # Ignorer si la conversion échoue

    # Sauvegarder le fichier Excel
    book.save(output_file)
    
    ws = book["Commodites"]

    # Initialiser les chaînes de caractères pour les colonnes C et E
    colCAmenities = ""
    colEAmenities = ""
    cellJ1Profitability = ""

    # Parcourir les lignes de la colonne C et E
    for i in range(3, 1000):
        cell_CAmenities = ws.cell(row=i, column=3)
        cell_EAmenities = ws.cell(row=i, column=5)
        cell_J1Profitability = ws.cell(row=i, column=4)

        if cell_CAmenities.value is not None:
            colCAmenities += str(cell_CAmenities.value) + ","

        if cell_EAmenities.value is not None:
            colEAmenities += str(cell_EAmenities.value) + ","

        if cell_J1Profitability.value is not None:
            cellJ1Profitability += str(cell_J1Profitability.value) + ","
    
    # Supprimer la dernière virgule
    colCAmenities = colCAmenities.rstrip(',')
    colEAmenities = colEAmenities.rstrip(',')
    cellJ1Profitability = cellJ1Profitability.rstrip(',')

    # Ajouter les validations de données
    dvcAmenities = DataValidation(type="list", formula1="=Reference!$M:$M", showDropDown=False)
    dveAmenities = DataValidation(type="list", formula1="=Reference!$E:$E", showDropDown=False)
    dvJ1Amenities = DataValidation(type="list", formula1="=Reference!$M:$M", showDropDown=False)

    ws.add_data_validation(dvcAmenities)
    ws.add_data_validation(dveAmenities)

    dvcAmenities.add("A3:A35")
    dveAmenities.add("B2:W2")

    
    # Ajouter la formule en anglais de B3 à W95
    for row in range(3, 96):
        for col in range(2, 24):  # B=2, W=23
            cell = ws.cell(row=row, column=col)
            cell.value = ('=IFERROR(SUMIFS(INDIRECT(_xlfn.XLOOKUP({0}2,Reference!$E:$E,Reference!$I:$I)), '
                        'INDIRECT(_xlfn.XLOOKUP({0}2,Reference!$E:$E,Reference!$H:$H)),Commodites!$A{1}), "")'
                        .format(chr(64 + col), row))
            

    ws = book["Rentabilite"]
    ws.add_data_validation(dvJ1Amenities)
    dvJ1Amenities.add("N1")
    
    ref_ws = book["Reference"]
    ws["N1"] = ref_ws["M2"].value

    if "Commodites1" in book.sheetnames:
        sheet = book["Commodites1"]
        book.remove(sheet)

    book.save(output_file)
    book.close()


if __name__ == "__main__":
    main()
