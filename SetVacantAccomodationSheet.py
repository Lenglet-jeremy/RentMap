import pandas as pd
import sys
from unidecode import unidecode
from openpyxl import load_workbook
import warnings
warnings.simplefilter("ignore", UserWarning)
# Ignorer les warnings spécifiques d'openpyxl
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl.styles.stylesheet")

print("Exécution du fichier : SetVancantAccomodationSheet.py")

if len(sys.argv) < 3:
    print("Usage: python SetVancantAccomodationSheet.py <output_filepath> <department>")
    sys.exit(1)

output_filepath = sys.argv[1]
department = sys.argv[2]

# Suppression du zéro en début de département si présent
if department.startswith("0"):
    department = department[1:]

# Charger le fichier source
input_filepath = "./data/logementVacant/LogementVacant.xlsx"
df = pd.read_excel(input_filepath, sheet_name=0)  # Charge la première feuille

# Vérifier que les colonnes nécessaires existent
if df.shape[1] < 3:
    print("Le fichier source ne contient pas suffisamment de colonnes.")
    sys.exit(1)

# Filtrage des lignes où la colonne A commence par le département suivi de 3 chiffres
df_filtered = df[df.iloc[:, 0].astype(str).str.match(rf"^{department}\d{{3}}")]

# Garder uniquement les lignes avec la valeur la plus haute dans la colonne C
latest_year = df_filtered.iloc[:, 2].max()
df_filtered = df_filtered[df_filtered.iloc[:, 2] == latest_year]

# Nettoyage de la colonne B
if df_filtered.shape[1] > 1:  # Vérifier si la colonne B existe
    df_filtered.iloc[:, 1] = (
        df_filtered.iloc[:, 1]
        .astype(str)
        .apply(unidecode)  # Supprime les accents
        .str.replace("'", " ")  # Remplace les apostrophes par des espaces
        .str.replace("-", " ")  # Remplace les tirets par des espaces
        .str.replace(r"\bSaint\b", "st", case=False, regex=True)  # Remplace "Saint" par "st"
    )

# Charger le fichier Excel existant
try:
    book = load_workbook(output_filepath)
    
    # Vérifier si la feuille "Habitat vacant" existe
    if "Habitat vacant" in book.sheetnames:
        sheet = book["Habitat vacant"]
        sheet.delete_rows(1, sheet.max_row)  # Supprimer tout le contenu de la feuille
    
    # Enregistrer les modifications avant d'écrire
    book.save(output_filepath)

    # Ouvrir en mode "a" (ajout) avec `ExcelWriter`
    with pd.ExcelWriter(output_filepath, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        df_filtered.to_excel(writer, sheet_name="Habitat vacant", index=False)

except FileNotFoundError:
    # Si le fichier n'existe pas, on le crée avec une seule feuille
    with pd.ExcelWriter(output_filepath, engine="openpyxl") as writer:
        df_filtered.to_excel(writer, sheet_name="Habitat vacant", index=False)

