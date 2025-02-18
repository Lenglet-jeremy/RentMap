import requests
from bs4 import BeautifulSoup
from random import uniform
import time
import sys
import os 
import pandas as pd
from openpyxl import load_workbook

# import SetRefinedSheet
import warnings
warnings.simplefilter("ignore", UserWarning)
# Ignorer les warnings spécifiques d'openpyxl
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl.styles.stylesheet")

url = "https://statistiques.pole-emploi.org/stmt/defm?lk=0&mm=0&pp=201607-&ss=1"

timeSleepDuration = uniform(2, 5)
output_file = sys.argv[1]

def timeSleep(min = 2, max = 5) : 
    print(f"TimeSleep de {timeSleepDuration}s")
    time.sleep(timeSleepDuration)

def checkPath(folderPath) : 
    if not os.path.exists(folderPath) : 
        os.makedirs(folderPath)

def getSource(url) : 
    timeSleep(timeSleepDuration)
    print(f"Scraping de l'url : {url}")
    response = requests.get(url)

    if response.status_code == 200 : 
        sourceCode = BeautifulSoup(response.text, "html.parser")
        sourceCode = sourceCode.prettify()

        checkPath("./data/unemployment/")

        with open("./data/unemployment/global.html", 'w', encoding='utf-8') as file: 
            file.write(sourceCode)

    else : 
        print("Erreur lors de la recupération des sources")

        
    extractUnemploymentFigures("./data/unemployment/global.html", output_file)




def extractUnemploymentFigures(filePath, outputFile):
    # Lire le contenu du fichier HTML
    with open(filePath, 'r', encoding='utf-8') as file:
        sourceCode = file.read()

    # Parser le HTML avec BeautifulSoup
    soup = BeautifulSoup(sourceCode, 'html.parser')

    # Trouver la première table avec la classe 'table tablesorter'
    table = soup.find('table', class_='table tablesorter')
    if not table:
        print("Table non trouvée.")
        return

    # Extraire les en-têtes de colonne (<th>) et limiter aux 2 premières colonnes
    thead = table.find('thead')
    th_values = [th.get_text(strip=True) for th in thead.find_all('th')][:2] if thead else []

    # Extraire les données des lignes (<tr> dans <tbody>)
    tbody = table.find('tbody')
    rows = []
    if tbody:
        for tr in tbody.find_all('tr'):
            row = [td.get_text(strip=True) for td in tr.find_all('td')][:2]  # Prendre uniquement les 2 premières colonnes
            rows.append(row)

    # Créer un DataFrame avec les 2 premières colonnes
    df_new = pd.DataFrame(rows, columns=th_values)

    
    df_new.iloc[:, 1] = df_new.iloc[:, 1].str.replace(r'\s+', '', regex=True)  # Supprime les espaces
    df_new.iloc[:, 1] = df_new.iloc[:, 1].str.replace(',', '.')  # Remplace les virgules par des points (si nécessaire)
    df_new.iloc[:, 1] = pd.to_numeric(df_new.iloc[:, 1], errors='coerce').fillna(0).astype(int)



    # Charger le fichier Excel existant
    try:
        wb = load_workbook(outputFile)
        ws = wb["Chomages"]  # Modifier si une autre feuille est concernée
    except FileNotFoundError:
        print("Fichier Excel non trouvé, création d'un nouveau.")
        wb = None

    if wb:
        # Supprimer les anciennes valeurs des deux premières colonnes
        max_row = ws.max_row
        for row in range(2, max_row + 1):  # Commence à 2 pour éviter l'en-tête
            ws.cell(row=row, column=1, value=None)
            ws.cell(row=row, column=2, value=None)

        # Sauvegarder après suppression
        wb.save(outputFile)

    # Réécriture des nouvelles valeurs sans toucher aux autres colonnes
    with pd.ExcelWriter(outputFile, mode='a', engine='openpyxl', if_sheet_exists='overlay') as writer:
        df_new.to_excel(writer, sheet_name='Chomages', index=False, startcol=0, startrow=0)


def export_to_excel(df, file_name):
    try:
        with pd.ExcelWriter(file_name, engine='xlsxwriter') as writer:
            df.to_excel(writer, index=False, sheet_name="Chomage globale")
    except Exception as e:
        print(f"Erreur lors de l'exportation vers Excel : {e}")


# ============================================================================================
# =========================================== main ===========================================
# ============================================================================================
def main() : 
    print("\n Execution du fichier : GetUnemployment.py")
    getSource(url)

if __name__ == "__main__" : 
    main()